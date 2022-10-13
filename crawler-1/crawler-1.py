
import requests
import openpyxl
from bs4 import BeautifulSoup
from time import sleep
import urllib3
urllib3.disable_warnings()

crawler_final_results = list(dict())


class Crawler():
    def __init__(self, _stockname):
        self.stockname = _stockname
        self.today_data = list(dict())
        self.fiveday_data = dict(list())
        self.craw_today_data()
        self.craw_fiveday_data()

    def craw_today_data(self):
        url = f'https://fubon-ebrokerdj.fbs.com.tw/z/zc/zco/zco_{self.stockname}_1.djhtm'
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:66.0) Gecko/20100101 Firefox/66.0",
            "Accept-Encoding": "*",
            "Connection": "keep-alive"
        }
        resp = requests.get(url, verify=True, headers=headers)
        soup = BeautifulSoup(resp.text, 'html5lib')
        brokers = soup.find_all('td', class_='t4t1')[:-4:2]
        today_amounts = soup.find_all('td', class_='t3n1')[2:-8:8]
        clinch_percentages = soup.find_all('td', class_='t3n1')[3:-8:8]
        today_totals = soup.find_all('td', class_='t3n1', colspan='4')[:2]
        for i in range(len(brokers)):
            try:
                if brokers[i].getText() == '\xa0':
                    continue
                self.today_data.append({
                    'broker': brokers[i].getText(),
                    'today_amount': int(today_amounts[i].getText().replace(',', '')),
                    'clinch_percentage': float(clinch_percentages[i].getText().replace('%', '')),
                    'today_all_buyin': int(today_totals[0].getText().replace(',', '')),
                    'today_all_sellout': int(today_totals[1].getText().replace(',', ''))
                })
            except:
                continue
                

    def craw_fiveday_data(self):
        url = f'https://fubon-ebrokerdj.fbs.com.tw/z/zc/zco/zco_{self.stockname}_2.djhtm'
        resp = requests.get(url, verify=False)
        soup = BeautifulSoup(resp.text, 'html5lib')
        brokers = soup.find_all('td', class_='t4t1')[:-4:2]
        #print(brokers)
        today_amounts = soup.find_all('td', class_='t3n1')[2:-8:8]
        today_totals = soup.find_all('td', class_='t3n1', colspan='4')[:2]
        for i in range(len(brokers)):
            try:
                if brokers[i].getText() == '\xa0':
                    continue
                self.fiveday_data[brokers[i].getText()] = [  # broker : today_amount today_all_buyin today_all_sellout
                    int(today_amounts[i].getText().replace(',', '')),
                    int(today_totals[0].getText().replace(',', '')),
                    int(today_totals[1].getText().replace(',', ''))
                ]
            except:
                continue


# stockname rank broker today_amount clinch_percentage fiveday_amount


def craw_stock(stockname, stockdata):
    crawler = Crawler(stockname)
    broke_data = iter(crawler.today_data)
    rankcounter = 1
    stock_result = list(dict())
    while rankcounter <= 15:
        now_broke_data = next(broke_data, -1)

        if now_broke_data == -1:
            break
        # print(now_broke_data)
        # print(crawler.fiveday_data[now_broke_data['broker']])

        if (now_broke_data['clinch_percentage'] > 3.0 and
                now_broke_data['today_amount'] / (0.01 * now_broke_data['clinch_percentage']) > 3000 and
                now_broke_data['broker'] in crawler.fiveday_data and
                crawler.fiveday_data[now_broke_data['broker']][0] > 2000 and
                crawler.fiveday_data[now_broke_data['broker']][0] > now_broke_data['today_amount'] * 1.3 and
                now_broke_data['today_amount'] > 300):

            stock_result.append({
                'stockname': stockname,
                'rank': rankcounter,
                'broker': now_broke_data['broker'],
                'today_amount': now_broke_data['today_amount'],
                'clinch_percentage': now_broke_data['clinch_percentage'],
                'fiveday_amount': crawler.fiveday_data[now_broke_data['broker']][0],
                'name': stockdata[0],
                'type': stockdata[1],
                'num': stockdata[2],
                'oneday_all_buyin': now_broke_data['today_all_buyin'],
                'oneday_all_sellout': now_broke_data['today_all_sellout'],
                'fiveday_all_buyin': crawler.fiveday_data[now_broke_data['broker']][1],
                'fiveday_all_sellout': crawler.fiveday_data[now_broke_data['broker']][2],
            })
            rankcounter += 1
    crawler_final_results.extend(stock_result)



def save_on_excel():
    workbook = openpyxl.load_workbook('stock_result.xlsx')
    sheet = workbook['ranking_sheet']
    cols = ['股票代號', '名稱', '類別', '排名', '買超券商', '近一日買超', '佔成交比重',
            '近五日買超', '數字', '單日合計買超張數', '單日合計賣超張數', '五日合計買超張數', '五日合計賣超張數']
    for idx, col in enumerate(cols):
        sheet.cell(
            row=1, column=idx + 1).value = col
    for r in range(2, sheet.max_row + 1):
        for c in range(len(cols)):
            sheet.cell(
                row=r, column=c + 1).value = ''

    for idx, crawler_final_result in enumerate(crawler_final_results):
        sheet.cell(
            row=idx + 2, column=1).value = crawler_final_result['stockname']
        sheet.cell(
            row=idx + 2, column=2).value = crawler_final_result['name']
        sheet.cell(
            row=idx + 2, column=3).value = crawler_final_result['type']
        sheet.cell(
            row=idx + 2, column=4).value = crawler_final_result['rank']
        sheet.cell(
            row=idx + 2, column=5).value = crawler_final_result['broker']
        sheet.cell(
            row=idx + 2, column=6).value = crawler_final_result['today_amount']
        sheet.cell(
            row=idx + 2, column=7).value = crawler_final_result['clinch_percentage']
        sheet.cell(
            row=idx + 2, column=8).value = crawler_final_result['fiveday_amount']
        sheet.cell(
            row=idx + 2, column=9).value = crawler_final_result['num']
        sheet.cell(
            row=idx + 2, column=10).value = crawler_final_result['oneday_all_buyin']
        sheet.cell(
            row=idx + 2, column=11).value = crawler_final_result['oneday_all_sellout']
        sheet.cell(
            row=idx + 2, column=12).value = crawler_final_result['fiveday_all_buyin']
        sheet.cell(
            row=idx + 2, column=13).value = crawler_final_result['fiveday_all_sellout']
    workbook.save('stock_result.xlsx')


def load_stockdata_on_excel():
    workbook = openpyxl.load_workbook('stock_list_test.xlsx')
    sheet = workbook['stock_list_test']
    stockdatas = dict()  # stockid : [name, type, num]
    for r in range(sheet.max_row):
        stockdatas[sheet.cell(row=r + 1, column=1).value] = [
            sheet.cell(row=r + 1, column=2).value,
            sheet.cell(row=r + 1, column=3).value,
            sheet.cell(row=r + 1, column=4).value
        ]
    return stockdatas


if __name__ == '__main__':
    stockdatas = load_stockdata_on_excel()
    for idx, stockid in enumerate(stockdatas):
        craw_stock(stockid, stockdatas[stockid])
        print(f'完成股票代號 {stockid} 爬取 | 當前進度 {idx+1}/{len(stockdatas)}')
        sleep(0.5)

    print(crawler_final_results)
    save_on_excel()
