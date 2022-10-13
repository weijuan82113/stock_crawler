
from importlib.machinery import all_suffixes
import requests
import openpyxl
from bs4 import BeautifulSoup
import pandas as pd
from time import sleep
import urllib3
urllib3.disable_warnings()

crawler_final_results = list(dict())

'''
crawler form
{
    'id' : int,
    'name' : str,
    'type' : str,
    'deal_price' : str,
    'one_in_out' : {
        '<dealer>' : [deal_over, deal_percent]
    },
    'five_in_out' : {
        '<dealer>' : deal_over
    },
    'total_in_over_one' : int,
    'total_out_over_one' : int,
    'total_in_over_five' : int,
    'total_out_over_five' : int,
    'max_PEratio' : int,
    'min_PEratio' : int,
    'performance' : [Q1, Q2, Q3, Q4],
    'capital' : int
}
'''

'''
result form
'''

all_result = []

class Crawler():
    def __init__(self, _stockid, _stockname, _stocktype):
        self.stockid = _stockid
        self.stockname = _stockname
        self.stocktype = _stocktype
        self.info_url = f'https://sjmain.esunsec.com.tw/z/zc/zca/zca_{self.stockid}.djhtm'
        self.performance_url = f'https://sjmain.esunsec.com.tw/z/zc/zcx/zcxD9Esunsec.djjs?A={self.stockid}'
        self.chips_url = f'https://sjmain.esunsec.com.tw/z/zc/zcx/zcxD1Esunsec.djjs?A={self.stockid}'
        self.one_in_out_url = f'https://sjmain.esunsec.com.tw/z/zc/zco/zco_{self.stockid}.djhtm'
        self.five_in_out_url = f'https://sjmain.esunsec.com.tw/z/zc/zco/zco_{self.stockid}_2.djhtm'

    def crawing(self):
        try:
            info_data = pd.read_html(requests.get(self.info_url).text)
            if len(info_data) == 1:
                return
            deal_price = info_data[2][7][1]
            mainInOut_table_one = pd.read_html(requests.get(self.one_in_out_url).text)[2]
            all_one_deal_over = dict()
            for i in range(6, len(mainInOut_table_one)-3):
                all_one_deal_over[mainInOut_table_one[0][i]] = [
                    mainInOut_table_one[3][i],
                    mainInOut_table_one[4][i]
                ]

            mainInOut_table_five = pd.read_html(requests.get(self.five_in_out_url).text)[2]
            all_five_deal_over = dict()
            for i in range(6, len(mainInOut_table_five)-3):
                all_five_deal_over[mainInOut_table_five[0][i]] = mainInOut_table_five[3][i]

            total_in_over_one = mainInOut_table_one[1][len(mainInOut_table_one)-3]
            total_out_over_one = mainInOut_table_one[6][len(mainInOut_table_one)-3]
            total_in_over_five = mainInOut_table_five[1][len(mainInOut_table_five)-3]
            total_out_over_five = mainInOut_table_five[6][len(mainInOut_table_five)-3]
            max_PEratio = info_data[2][3][4]
            min_PEratio = info_data[2][5][4]
            performance_table = pd.read_html(requests.get(self.performance_url).text)[0][3]
            performances = [performance_table[i] for i in range(2,6)]
            capital_table = pd.read_html(requests.get(self.chips_url).text)[0]
            capital = float(capital_table[1][2])/(float(capital_table[2][2].strip('%'))/100) if float(capital_table[2][2].strip('%'))!=0 else ''
            return {
                'id' : self.stockid,
                'name' : self.stockname,
                'type' : self.stocktype,
                'deal_price' : deal_price,
                'one_in_out' : all_one_deal_over,
                'five_in_out' : all_five_deal_over,
                'total_in_over_one' : total_in_over_one,
                'total_out_over_one' : total_out_over_one,
                'total_in_over_five' : total_in_over_five,
                'total_out_over_five' : total_out_over_five,
                'max_PEratio' : max_PEratio,
                'min_PEratio' : min_PEratio,
                'performance' : performances,
                'capital' : capital
            }
        except Exception as err:
            pass

class Analyzer():
    def __init__(self, _stockid, _stockname, _stocktype):
        crawler = Crawler(_stockid, _stockname, _stocktype)
        self.stockid = _stockid
        self.stockname = _stockname
        self.stocktype = _stocktype
        self.crawler_data = crawler.crawing()

    def analyzing(self):
        if not self.crawler_data:
            return
        for one_in_out_dealer in self.crawler_data['one_in_out']:
            if float(self.crawler_data['one_in_out'][one_in_out_dealer][0])*float(self.crawler_data['deal_price']) > 20000\
                and float(self.crawler_data['deal_price']) <= 300\
                and one_in_out_dealer in self.crawler_data['five_in_out']\
                and float(self.crawler_data['five_in_out'][one_in_out_dealer])>float(self.crawler_data['one_in_out'][one_in_out_dealer][0])*1.33:
                all_result.append({
                    'id' : self.stockid,
                    'name' : self.stockname,
                    'type' : self.stocktype,
                    'deal_price' : self.crawler_data['deal_price'],
                    'dealer_name' : one_in_out_dealer,
                    'one_in_mount' : self.crawler_data['one_in_out'][one_in_out_dealer][0],
                    'one_in_percent' : self.crawler_data['one_in_out'][one_in_out_dealer][1],
                    'five_in_mount' : self.crawler_data['five_in_out'][one_in_out_dealer],
                    'total_in_over_one' : self.crawler_data['total_in_over_one'],
                    'total_out_over_one' : self.crawler_data['total_out_over_one'],
                    'total_in_over_five' : self.crawler_data['total_in_over_five'],
                    'total_out_over_five' : self.crawler_data['total_out_over_five'],
                    'max_PEratio' : self.crawler_data['max_PEratio'],
                    'min_PEratio' : self.crawler_data['min_PEratio'],
                    'performance' : self.crawler_data['performance'],
                    'capital' : self.crawler_data['capital']
                })

def save_on_excel():
    workbook = openpyxl.load_workbook('stock_result.xlsx')
    sheet = workbook['ranking_sheet']
    cols = ['股票代號', '名稱', '類別', '成交價', '買超券商', '近一日買超', '佔成交比重',
            '近五日買超', '單日合計買超張數', '單日合計賣超張數', '五日合計買超張數', '五日合計賣超張數',
            '最高本益比', '最低本益比', '經營績效1', '經營績效2', '經營績效3', '經營績效4','股本']
    for idx, col in enumerate(cols):
        sheet.cell(
            row=1, column=idx + 1).value = col
    for r in range(2, sheet.max_row + 1):
        for c in range(len(cols)):
            sheet.cell(
                row=r, column=c + 1).value = ''

    for idx, stock_details in enumerate(all_result):
        sheet.cell(
            row=idx + 2, column=1).value = stock_details['id']
        sheet.cell(
            row=idx + 2, column=2).value = stock_details['name']
        sheet.cell(
            row=idx + 2, column=3).value = stock_details['type']
        sheet.cell(
            row=idx + 2, column=4).value = stock_details['deal_price']
        sheet.cell(
            row=idx + 2, column=5).value = stock_details['dealer_name']
        sheet.cell(
            row=idx + 2, column=6).value = stock_details['one_in_mount']
        sheet.cell(
            row=idx + 2, column=7).value = stock_details['one_in_percent']
        sheet.cell(
            row=idx + 2, column=8).value = stock_details['five_in_mount']
        sheet.cell(
            row=idx + 2, column=9).value = stock_details['total_in_over_one']
        sheet.cell(
            row=idx + 2, column=10).value = stock_details['total_out_over_one']
        sheet.cell(
            row=idx + 2, column=11).value = stock_details['total_in_over_five']
        sheet.cell(
            row=idx + 2, column=12).value = stock_details['total_out_over_five']
        sheet.cell(
            row=idx + 2, column=13).value = stock_details['max_PEratio']
        sheet.cell(
            row=idx + 2, column=14).value = stock_details['min_PEratio']
        sheet.cell(
            row=idx + 2, column=15).value = stock_details['performance'][0]
        sheet.cell(
            row=idx + 2, column=16).value = stock_details['performance'][1]
        sheet.cell(
            row=idx + 2, column=17).value = stock_details['performance'][2]
        sheet.cell(
            row=idx + 2, column=18).value = stock_details['performance'][3]
        sheet.cell(
            row=idx + 2, column=19).value = stock_details['capital']
    workbook.save('stock_result.xlsx')

def load_stockdata_on_excel():
    workbook = openpyxl.load_workbook('stock_list.xlsx')
    sheet = workbook['stock_list_test']
    stockdatas = dict()  # stockid : [name, type]
    for r in range(sheet.max_row):
        stockdatas[sheet.cell(row=r + 1, column=1).value] = [
            sheet.cell(row=r + 1, column=2).value.strip(),
            sheet.cell(row=r + 1, column=3).value.strip(),
        ]
    return stockdatas

if __name__ == '__main__':
    stockdatas = load_stockdata_on_excel()
    for idx, stockid in enumerate(stockdatas):
        analyzer = Analyzer(stockid, stockdatas[stockid][0], stockdatas[stockid][1])
        analyzer.analyzing()
        print(f'[+] 完成股票代號 {stockid} {stockdatas[stockid][0]} {stockdatas[stockid][1]} 爬取 | 當前進度 {idx+1}/{len(stockdatas)}')
    # print("Output:", all_result)
    save_on_excel()
    pass
